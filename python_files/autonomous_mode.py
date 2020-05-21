import cv2
import numpy as np
import requests
import httpx
import asyncio
#para excel
from openpyxl import load_workbook
from openpyxl import Workbook
import os.path
import traceback
import keras
from keras.models import Sequential
from keras.models import model_from_json
import copy
from keras.models import load_model

url="http://192.168.137.199:81/stream"
url2="http://192.168.137.199:80"

#----------  load the model------------------#
# load json and create model
json_file = open("C:/Users/ASUS/Desktop/Proyecto Carrito/modelA.json", 'r')
loaded_model_json = json_file.read()
json_file.close()
loaded_model = model_from_json(loaded_model_json)
# load weights into new model
loaded_model.load_weights("C:/Users/ASUS/Desktop/Proyecto Carrito/model_carritoA.h5")
print("Loaded model from disk")

model_signs = load_model('C:/Users/ASUS/Desktop/Proyecto Carrito/model_signs_3.h5')
#---------- ----------------------------------------#
def display_lines(image, lines):
    line_image = np.zeros_like(image)   #we make the pixels black
    if lines is not None:
        for line in lines:
            for x1, y1, x2, y2 in line:
                 cv2.line(line_image, (x1, y1), (x2, y2), (255, 0, 0), 10)#print(line) #coordenadas de las lineas

    return line_image

def region_of_interest(image):
    height = image.shape[0]
    polygons = np.array([[(200, height), (650, height), (650, 190), (200, 190)]], np.int32)
    #mask = np.zeros_like(image)
    cv2.fillPoly(image, polygons, 0) #0 = color black, 255 = color white
    #masked_image = cv2.bitwise_and(image, mask)
    return image

def canny(image):
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    kernel = 5
    blur = cv2.GaussianBlur(gray, (kernel, kernel), 0)
    canny = cv2.Canny(blur, 50, 150)
    return canny

def img_preprocess(img):
    sensitivity=110
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    lower_white = np.array([0,0,255-sensitivity])
    upper_white = np.array([255,sensitivity,255])
    mask = cv2.inRange(hsv, lower_white, upper_white)
    img = cv2.resize(mask, (100, 100)) #we expand the image
    img = np.stack((img,)*3, axis=-1)
    img = img/255 #normaize the image
    return img

def grayscale(img):
    imagen = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    return imagen

def equalize(img):
    img = cv2.equalizeHist(img)
    return img

def preprocessing(img):
    img = grayscale(img)
    img = equalize(img)
    #another prepocesing technique called normalization that consist in divide all the pixels intensities of our image by 255
    img = img / 255
    return img

def simbolo_traf(img):
    simbolo = -1
    gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    gray=cv2.GaussianBlur(gray,(3,3),0)
    edged=cv2.Canny(gray,50,150) #deteccion bordes
    kernel=cv2.getStructuringElement(cv2.MORPH_RECT,(5,5))
    closed=cv2.morphologyEx(edged,cv2.MORPH_CLOSE,kernel,iterations=2)

    cnts,_=cv2.findContours(closed.copy(), cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE) #encontrar contornos
    #total=0
    for c in cnts:
        area=cv2.contourArea(c)
        #print "area",area
        if area>1700:
        #aproximacion de contorno
            peri=cv2.arcLength(c,True) #Perimetro
            approx=cv2.approxPolyDP(c,0.02*peri,True)

            #Si la aproximacion tiene 4 vertices correspondera a un rectangulo (Libro)
            if len(approx)==4:
                cv2.drawContours(img,[approx],-1,(0,255,0),3,cv2.LINE_AA)
                print(approx, end = '')
                x, y, w, h = cv2.boundingRect(approx)
                v1, v2, v3, v4 = approx
                tri_1 = [v1, v2, v3]
                tri_2 = [v1,v2, v4]
                A1 = cv2.contourArea(np.float32(tri_1))
                A2 = cv2.contourArea(np.float32(tri_2))
                aspect_ratio = float(w)/h
                print (aspect_ratio)
                if  1.0 < aspect_ratio < 1.6 and abs(A1-A2) < 500:
                    print(A1, A2)
                    out = img[y:y+h, x:x+w]

                    #Preprocess image
                    out = np.asarray(out)
                    out = cv2.resize(out, (32, 32))
                    out = preprocessing(out)
                    print(out.shape)
                    out = out.reshape(1, 32, 32, 1)
                    #Test image
                    simbolo = int(model_signs.predict_classes(out))
                    print("predicted sign: "+ str(simbolo))

                    # Show the output image
                    #cv2.imwrite("C:/Users/ASUS/Desktop/simb/im%d.jpg" %(area), out)
                #total+=1
    #Poner texto en imagen
    #letrero= 'Objetos: '+ str(total)
    #cv2.putText(img,letrero,(10,150),cv2.FONT_HERSHEY_SIMPLEX,1,(255,0,0),2)
    #Mostramos imagen
    #cv2.imshow("video", img)
    return simbolo


async def control(url_get,params):
    async with httpx.AsyncClient() as client:
        await client.get(url2+url_get+"?"+params)

class Velocidades():
    vel_max_esp="235"
    vel_vuelta_esp="220"
    giro = 0.5
    def get_velocidad_esp(self,velocidad):
        if velocidad > self.giro:
            return self.vel_vuelta_esp+"000"
        elif velocidad >= -self.giro and velocidad <= self.giro:
            return self.vel_max_esp+self.vel_max_esp
        elif velocidad < -self.giro:
            return "000"+self.vel_vuelta_esp

    def set_velocidad_max(self,velocidad):
        self.vel_max_esp=velocidad
    def set_velocidad_vuelta(self,velocidad):
        self.vel_vuelta_esp=velocidad

vel = Velocidades()

cont = 0
led_sent=""
last_vel=""
vel_sent=""
img_cpy=""
imagenes=[]
opc = input("Guardar imagenes? S/N")
opc = opc.lower()

loop = asyncio.get_event_loop()
loop.run_until_complete(control('/direccion',"w"))

CAMERA_BUFFRER_SIZE=1024
stream=requests.get(url, verify=False, stream=True)
bts=b''

while True:
    try:
        bts+=stream.raw.read(CAMERA_BUFFRER_SIZE)
        jpghead=bts.find(b'\xff\xd8')
        jpgend=bts.find(b'\xff\xd9')
        if jpghead>-1 and jpgend>-1:
            jpg=bts[jpghead:jpgend+2]
            bts=bts[jpgend+2:]
            img=cv2.imdecode(np.frombuffer(jpg,dtype=np.uint8),cv2.IMREAD_UNCHANGED)
            #img=cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
            v=cv2.flip(img,0)
            h=cv2.flip(img,1)
            p=cv2.flip(img,-1)

            frame=p
            h,w=frame.shape[:2]
            #Imagen
            img=cv2.resize(frame,(800,600))  #we get the image from the Server (ESP32-CAM)
            k=cv2.waitKey(10)
            if k==ord('q'):
                loop = asyncio.get_event_loop()
                loop.run_until_complete(control('/vel',"000000"))
                #peticion('/vel',"000000")
                break
            '''
            #-----------activacion de color verde ---------------#
            if img_green(img):
                if led_sent!="l":
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(control('/led','l'))
                    led_sent="l"
            else:
                if led_sent!="o":
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(control('/led','o'))
                    led_sent="o"
            #------------------------------------- ---------------#
            '''
            simb = simbolo_traf(img)
            if simb==0:  #baja velocidad
                vel.set_velocidad_max("235")
                vel.set_velocidad_vuelta("220")
            if simb==1:  #alta velocidad
                vel.set_velocidad_max("255")
                vel.set_velocidad_vuelta("220")
            if simb==2:  #prender led (rombo)
                if led_sent!="l":
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(control('/led','l'))
                    led_sent="l"
            if simb==3:
                vel.set_velocidad_max("000")
                vel.set_velocidad_vuelta("000")
            if simb==4:
                if led_sent!="o":
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(control('/led','o'))
                    led_sent="o"

            img_cpy=copy.copy(img)
            img = img_preprocess(img)   #we preprocess the image before put in the neural network
            cv2.imshow("Carrito",img)       #we show the image
            img = img.reshape(1, 100, 100, 3)
            prediccion = loaded_model.predict(img)
            #prediccion+=1.5
            velocidad=vel.get_velocidad_esp(prediccion)
            print("Velocidad esp: %s Prediccion: %f" %(velocidad,prediccion))

            if last_vel == velocidad:
                cont+= 1

            if cont >= 2:
                cont = 0
                if velocidad!=vel_sent:
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(control('/vel',velocidad))
                    #peticion('/vel',velocidad)
                    vel_sent=velocidad
            last_vel = velocidad

            if opc=="s":
                imagenes.append([img_cpy,prediccion,velocidad])
    except:
        traceback.print_exc()

ruta="C:/Users/ASUS/Desktop/Proyecto Carrito"
def get_filas():
    cont=0
    for row in ws.values:
        cont+=1
    return cont
def agregar_fila(imagen,prediccion,velocidad):
    if get_filas()==0:
        ws["A1"]=imagen
        ws["B1"]=prediccion
        ws["C1"]=velocidad
    else:
        fila=get_filas()+1
        ws.cell(row=fila, column=1, value=imagen)
        ws.cell(row=fila, column=2, value=prediccion)
        ws.cell(row=fila, column=3, value=velocidad)

opc = input("Guardar imagenes? S/N")
opc = opc.lower()
if(opc=="s"):
    excel = Workbook()
    ws = excel.active
    cont=0
    for imagen,prediccion,velocidad in imagenes:
        print(cont)
        r=ruta+"/imag/imagen%d.jpg" %(cont)
        cv2.imwrite(r,imagen)
        agregar_fila("imagen%d.jpg" %(cont),str(prediccion),str(velocidad))
        cont+=1
    guardando=True
    while guardando:
        try:
            print("Guardando...")
            excel.save(ruta+"/exele/excel.xlsx")
            guardando=False
        except:
            traceback.print_exc()
            opc= input("Intentar de nuevo? S/N ")
            if (opc.lower()=="n"):
                guardando=False

'''
async def control(url_get):
    async with httpx.AsyncClient() as client:
        await client.get(url2+url_get)
'''
